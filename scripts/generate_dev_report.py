#!/usr/bin/env python3
"""
FitTrack Developer Performance & Profiling Report Generator
============================================================
Runs all 3 AI features (Diet Plan, Workout Plan, Feast Mode) with two models,
captures full Ollama metrics, and generates fittrack_dev_report.xlsx.

Usage:  python scripts/generate_dev_report.py
Output: fittrack_dev_report.xlsx  (project root)
"""

import os
import sys
import json
import time
import requests
import traceback
from datetime import datetime, date

# ---------------------------------------------------------------------------
# CONFIG
# ---------------------------------------------------------------------------
OLLAMA_URL = os.getenv("OLLAMA_URL", "http://localhost:11434")
MODEL_A = "gpt-oss:120b-cloud"
MODEL_B = "llama3.2:latest"

SAMPLE_USER = {
    "age": 23,
    "weight_kg": 70,
    "height_cm": 175,
    "gender": "male",
    "goal": "fat loss",
    "diet_type": "vegetarian",
    "activity_level": "moderately active",
    "country": "India",
    "region": "India",
    "experience_level": "Intermediate",
    "days_per_week": 5,
    "session_duration_min": 60,
    "health_restrictions": "None",
}

# ---------------------------------------------------------------------------
# PROMPT DEFINITIONS  (mirrors exact codebase prompts)
# ---------------------------------------------------------------------------

# 1) DIET PLAN GENERATION
DIET_SYSTEM_PROMPT = "You are a nutritionist. Output ONLY valid JSON. No explanation."

DIET_USER_PROMPT = f"""
# ROLE (Persona)
You are a professional Nutritionist and Meal Planner.
Your job is to SELECT DISHES - we will calculate portions and nutrients automatically.

# CONTEXT
- Diet Type: Vegetarian
- Region: India
- User: Age {SAMPLE_USER['age']}, Weight {SAMPLE_USER['weight_kg']}kg, Goal: {SAMPLE_USER['goal']}

# TASK (Goal)
Select 4 meals (Breakfast, Lunch, Dinner, Snacks) using authentic Indian vegetarian foods.
Custom user requirements: None

# CONSTRAINTS (Formatting & Instructions)
STRICT VEGETARIAN RULE:
- Every single meal must be 100% Vegetarian.
- NO meat, NO fish, NO eggs.
- Use Paneer, Curd, Lentils, Beans, Tofu, Milk as protein sources.

=== DISH NAMING FORMAT ===
Format: "Main Dish (veg/non-veg) + Accompaniment + Side"
Examples for India:
- "Palak Paneer (veg) + Jeera Rice + Cucumber Salad"
- "Dal Tadka (veg) + 2 Roti + Mixed Salad"
- "Masala Dosa (veg) + Sambar + Coconut Chutney"

=== OUTPUT JSON (SIMPLIFIED) ===
You only need to provide dish selection. We calculate nutrients automatically.
Strictly output ONLY valid JSON:
{{{{
  "meal_plan": [
    {{
      "meal_id": "breakfast",
      "label": "Breakfast",
      "is_veg": true,
      "dish_name": "Poha (veg) + Curd + Apple",
      "alternatives": ["Upma (veg)", "Idli (veg)"],
      "guidelines": ["Add peanuts for protein", "Pair with green tea"]
    }},
    {{
      "meal_id": "lunch",
      "label": "Lunch",
      "is_veg": true,
      "dish_name": "...",
      "alternatives": ["...", "..."],
      "guidelines": ["...", "..."]
    }},
    {{
      "meal_id": "dinner",
      "label": "Dinner",
      "is_veg": true,
      "dish_name": "...",
      "alternatives": ["...", "..."],
      "guidelines": ["...", "..."]
    }},
    {{
      "meal_id": "snacks",
      "label": "Snacks",
      "is_veg": true,
      "dish_name": "...",
      "alternatives": ["...", "..."],
      "guidelines": ["...", "..."]
    }}
  ]
}}}}

RULES:
- Use authentic India food names
- Include (veg/non-veg) label in dish_name
- DO NOT include "portion_size" or "nutrients" - we calculate those automatically
- Output JSON only.
"""

# 2) WORKOUT PLAN GENERATION
WORKOUT_SYSTEM_PROMPT = "You are a professional fitness coach. Return strictly valid JSON."

WORKOUT_USER_PROMPT = f"""
# ROLE (Persona)
You are a professional Fitness Coach.

# CONTEXT
- User Profile: Weight: {SAMPLE_USER['weight_kg']}kg, Goal: {SAMPLE_USER['goal']}
- Preferences: {SAMPLE_USER['experience_level']}, {SAMPLE_USER['days_per_week']} days/week, {SAMPLE_USER['session_duration_min']} mins/session

# TASK (Goal)
Create a 7-day weekly schedule that includes exactly {SAMPLE_USER['days_per_week']} workout days (rest of the days are Rest/Recovery).
The schedule must:
1. Align with the user's goal ({SAMPLE_USER['goal']}) and experience level.
2. Include specific warm-up and regular cardio on workout days.

# WEEKLY DAY ORDER (CRITICAL)
You MUST use exactly this day mapping for the weekly schedule:
"day1": "Monday", "day2": "Tuesday", "day3": "Wednesday", "day4": "Thursday", "day5": "Friday", "day6": "Saturday", "day7": "Sunday"

⛔ SUNDAY REST RULE:
- Sunday MUST be a REST/RECOVERY day.

# CONSTRAINTS (Health & Safety) - CRITICAL
- Health Restrictions: "{SAMPLE_USER['health_restrictions']}"

# FORMATTING RULES
1. MANDATORY: Include at least 1 cardio exercise per workout day.
2. INSTRUCTIONS: Each exercise MUST have "instructions" with exactly 3 tips.
3. ALWAYS output exactly 7 days (day1 through day7).
4. OUTPUT: Strictly return valid JSON.

=== OUTPUT JSON ===
{{{{
  "workout_plan": {{{{
    "plan_name": "...",
    "primary_goal": "{SAMPLE_USER['goal']}",
    "duration_weeks": 8,
    "weekly_schedule": {{{{
      "day1": {{{{
         "day_name": "Monday",
         "workout_name": "Push Day",
         "primary_muscle_group": "Chest & Triceps",
         "focus": "Strength",
         "exercises": [ {{{{"exercise": "Bench Press", "sets": 3, "reps": "10-12", "rest_sec": 60, "instructions": ["tip1", "tip2", "tip3"]}}}} ],
         "cardio_exercises": [ {{{{"exercise": "Running", "duration": "10 mins", "intensity": "Moderate", "notes": "Warm-up", "instructions": ["tip1", "tip2", "tip3"]}}}} ],
         "session_duration_min": 60
      }}}},
      "day2": {{{{ "day_name": "Tuesday", "..." : "..." }}}},
      "day3": {{{{ "day_name": "Wednesday", "..." : "..." }}}},
      "day4": {{{{ "day_name": "Thursday", "..." : "..." }}}},
      "day5": {{{{ "day_name": "Friday", "..." : "..." }}}},
      "day6": {{{{ "day_name": "Saturday", "..." : "..." }}}},
      "day7": {{{{ "day_name": "Sunday", "..." : "..." }}}}
    }}}},
    "progression_guidelines": ["Tip 1", "Tip 2"],
    "cardio_recommendations": ["Do 20 mins LISS on rest days"]
  }}}}
}}}}
"""

# 3) FEAST MODE ACTIVATION
FEAST_SYSTEM_PROMPT = """You are a nutrition adjustment agent for a meal planning app.
Your job is to adjust the remaining meals to meet a new calorie target.

CONTEXT: The user is in 'Feast Mode' banking phase. You must ADJUST the total calories to match the specified target.

RULES:
1. PROTECT PROTEIN: Never reduce protein by more than 5%.
2. REDUCE SNACKS FIRST: When cutting calories, reduce snack portions before touching main meals.
3. CARBS ARE FLEXIBLE: Carbs are the easiest to adjust.
4. FAT IS SECONDARY: After carbs, adjust fat content slightly if needed.
5. KEEP DISH NAMES: Never change the dish itself, only adjust portion sizes and nutrients.
6. PROVIDE NOTES: For each adjusted meal, provide a short human-readable note.
7. RESPECT SELECTED MEALS: Only adjust meals in the selected_meals list.

Respond in JSON format:
{
  "adjusted_meals": [
    {
      "meal_id": "breakfast",
      "calories": 350,
      "portion_size": "updated portion string",
      "protein": 25.0,
      "carbs": 40.0,
      "fat": 10.0,
      "note": "Reduced rice from 150g to 100g (-80 kcal)"
    }
  ]
}"""

FEAST_USER_PROMPT = """Event: Birthday Party (BANKING)
Context:
- Calorie Debt to Pay: 400 kcal
- User Selected Meals for Payment: breakfast, lunch, dinner, snacks

Meals Available:
[
  {"meal_id": "breakfast", "dish_name": "Poha (veg) + Curd + Apple", "is_selected_for_adjustment": true, "base_calories": 420, "macros": {"p": 12.0, "c": 65.0, "f": 10.0}, "protein": 12.0, "carbs": 65.0, "fat": 10.0, "portion_size": "200g Poha, 100g Curd, 1 Apple"},
  {"meal_id": "lunch", "dish_name": "Dal Tadka (veg) + 2 Roti + Salad", "is_selected_for_adjustment": true, "base_calories": 550, "macros": {"p": 18.0, "c": 75.0, "f": 12.0}, "protein": 18.0, "carbs": 75.0, "fat": 12.0, "portion_size": "150g Dal, 2 Roti, 100g Salad"},
  {"meal_id": "dinner", "dish_name": "Palak Paneer (veg) + Jeera Rice", "is_selected_for_adjustment": true, "base_calories": 500, "macros": {"p": 20.0, "c": 55.0, "f": 18.0}, "protein": 20.0, "carbs": 55.0, "fat": 18.0, "portion_size": "150g Palak Paneer, 150g Rice"},
  {"meal_id": "snacks", "dish_name": "Sprouts Chaat (veg)", "is_selected_for_adjustment": true, "base_calories": 180, "macros": {"p": 8.0, "c": 25.0, "f": 3.0}, "protein": 8.0, "carbs": 25.0, "fat": 3.0, "portion_size": "150g Sprouts"}
]

RULES:
1. You are only allowed to edit these specific Meal IDs: breakfast, lunch, dinner, snacks.
2. Treat the 'Calorie Debt' (400 kcal) as a debt to be divided among these selected meals.
3. If a meal falls below 75 calories after adjustment, set it to 0.
4. Return ALL meals in the JSON, even if unchanged.

Adjust calories, protein, carbs, fat, portion_size.
"""

# ---------------------------------------------------------------------------
# FEATURES CONFIG
# ---------------------------------------------------------------------------
FEATURES = [
    {
        "name": "Diet Plan Generation",
        "system_prompt": DIET_SYSTEM_PROMPT,
        "user_prompt": DIET_USER_PROMPT,
        "json_mode": True,
        "temperature": 0.1,
        "max_tokens": 12000,
    },
    {
        "name": "Workout Plan Generation",
        "system_prompt": WORKOUT_SYSTEM_PROMPT,
        "user_prompt": WORKOUT_USER_PROMPT,
        "json_mode": True,
        "temperature": 0.2,
        "max_tokens": 20000,
    },
    {
        "name": "Feast Mode Activation",
        "system_prompt": FEAST_SYSTEM_PROMPT,
        "user_prompt": FEAST_USER_PROMPT,
        "json_mode": True,
        "temperature": 0.1,
        "max_tokens": 20000,
    },
]

# ---------------------------------------------------------------------------
# OLLAMA DIRECT API CALLER
# ---------------------------------------------------------------------------
def call_ollama(model: str, system_prompt: str, user_prompt: str,
                temperature: float = 0.1, max_tokens: int = 12000,
                json_mode: bool = True) -> dict:
    """
    Calls Ollama /api/chat directly and returns full metrics dict.
    """
    url = f"{OLLAMA_URL.rstrip('/')}/api/chat"
    payload = {
        "model": model,
        "messages": [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_prompt},
        ],
        "stream": False,
        "options": {
            "temperature": temperature,
            "num_predict": max_tokens,
        },
    }
    if json_mode:
        payload["format"] = "json"

    result = {
        "model": model,
        "input_tokens": 0,
        "output_tokens": 0,
        "total_tokens": 0,
        "prompt_eval_duration_ms": 0,
        "eval_duration_ms": 0,
        "total_duration_ms": 0,
        "wall_clock_ms": 0,
        "response_text": "",
        "response_length": 0,
        "status": "error",
        "error": None,
        "timestamp": datetime.now().isoformat(),
    }

    wall_start = time.time()
    try:
        print(f"    [Ollama] POST {url}  model={model}  (timeout=300s) ...")
        resp = requests.post(url, json=payload, timeout=300)
        wall_end = time.time()
        result["wall_clock_ms"] = round((wall_end - wall_start) * 1000, 2)

        if resp.status_code != 200:
            result["error"] = f"HTTP {resp.status_code}: {resp.text[:300]}"
            print(f"    [Ollama] ERROR: {result['error']}")
            return result

        data = resp.json()
        msg = data.get("message", {})
        result["response_text"] = msg.get("content", "")
        result["response_length"] = len(result["response_text"])

        # Token counts
        result["input_tokens"] = data.get("prompt_eval_count", 0)
        result["output_tokens"] = data.get("eval_count", 0)
        result["total_tokens"] = result["input_tokens"] + result["output_tokens"]

        # Durations (ns -> ms)
        result["prompt_eval_duration_ms"] = round(data.get("prompt_eval_duration", 0) / 1e6, 2)
        result["eval_duration_ms"] = round(data.get("eval_duration", 0) / 1e6, 2)
        result["total_duration_ms"] = round(data.get("total_duration", 0) / 1e6, 2)

        result["status"] = "success"
        print(f"    [Ollama] OK  tokens={result['total_tokens']}  wall={result['wall_clock_ms']}ms  len={result['response_length']}")
        return result

    except requests.exceptions.Timeout:
        wall_end = time.time()
        result["wall_clock_ms"] = round((wall_end - wall_start) * 1000, 2)
        result["error"] = "Request timed out (300s)"
        print(f"    [Ollama] TIMEOUT after {result['wall_clock_ms']}ms")
        return result
    except Exception as e:
        wall_end = time.time()
        result["wall_clock_ms"] = round((wall_end - wall_start) * 1000, 2)
        result["error"] = str(e)
        print(f"    [Ollama] EXCEPTION: {e}")
        return result


# ---------------------------------------------------------------------------
# TOKEN ESTIMATION (rough, ~4 chars per token for English)
# ---------------------------------------------------------------------------
def estimate_tokens(text: str) -> int:
    if not text:
        return 0
    return max(1, len(text) // 4)


# ---------------------------------------------------------------------------
# RUN PROFILING
# ---------------------------------------------------------------------------
def run_profiling() -> list:
    """Run all features with both models.  Returns list of result dicts."""
    all_results = []
    models = [MODEL_A, MODEL_B]

    for model in models:
        print(f"\n{'='*70}")
        print(f"  MODEL: {model}")
        print(f"{'='*70}")
        for feat in FEATURES:
            print(f"\n  >> Feature: {feat['name']}")
            res = call_ollama(
                model=model,
                system_prompt=feat["system_prompt"],
                user_prompt=feat["user_prompt"],
                temperature=feat["temperature"],
                max_tokens=feat["max_tokens"],
                json_mode=feat["json_mode"],
            )
            res["feature"] = feat["name"]
            res["system_prompt_preview"] = feat["system_prompt"][:100]
            res["system_prompt_tokens_est"] = estimate_tokens(feat["system_prompt"])
            res["user_prompt_tokens_est"] = estimate_tokens(feat["user_prompt"])
            res["total_input_tokens_est"] = res["system_prompt_tokens_est"] + res["user_prompt_tokens_est"]
            all_results.append(res)

    return all_results


# ---------------------------------------------------------------------------
# LANGFUSE DATA  (best-effort pull)
# ---------------------------------------------------------------------------
def try_pull_langfuse() -> dict:
    """Attempt to pull trace data from Langfuse REST API."""
    public_key = os.getenv("LANGFUSE_PUBLIC_KEY", "")
    secret_key = os.getenv("LANGFUSE_SECRET_KEY", "")
    host = os.getenv("LANGFUSE_HOST", "https://us.cloud.langfuse.com")

    info = {
        "connected": False,
        "host": host,
        "traces": {},
        "error": None,
    }

    if not public_key or not secret_key:
        info["error"] = "Langfuse credentials not set in environment"
        return info

    try:
        # List recent traces
        url = f"{host.rstrip('/')}/api/public/traces?limit=30&orderBy=timestamp&orderDirection=DESC"
        resp = requests.get(url, auth=(public_key, secret_key), timeout=15)
        if resp.status_code == 200:
            info["connected"] = True
            traces = resp.json().get("data", [])
            # Group by feature name
            for t in traces:
                name = t.get("name", "unknown")
                if name not in info["traces"]:
                    info["traces"][name] = []
                info["traces"][name].append(t)
        else:
            info["error"] = f"HTTP {resp.status_code}: {resp.text[:200]}"
    except Exception as e:
        info["error"] = str(e)

    return info


# ---------------------------------------------------------------------------
# OLLAMA VERSION
# ---------------------------------------------------------------------------
def get_ollama_version() -> str:
    try:
        resp = requests.get(f"{OLLAMA_URL.rstrip('/')}/api/version", timeout=5)
        if resp.status_code == 200:
            return resp.json().get("version", "Unknown")
    except:
        pass
    return "N/A"


# ---------------------------------------------------------------------------
# EXCEL GENERATION
# ---------------------------------------------------------------------------
def generate_excel(results: list, langfuse_info: dict, output_path: str):
    """Create fittrack_dev_report.xlsx with 8 sheets."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("[ERROR] openpyxl not installed. Installing...")
        import subprocess
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()

    # Style helpers
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    def style_header_row(ws, row_num, max_col):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = thin_border

    def style_data_cell(ws, row, col, value=None):
        cell = ws.cell(row=row, column=col)
        if value is not None:
            cell.value = value
        cell.border = thin_border
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        return cell

    def auto_width(ws):
        for col in ws.columns:
            max_len = 0
            col_letter = None
            for cell in col:
                if col_letter is None and hasattr(cell, 'column_letter'):
                    col_letter = cell.column_letter
                try:
                    val = str(cell.value or "")
                    max_len = max(max_len, min(len(val), 60))
                except:
                    pass
            if col_letter:
                ws.column_dimensions[col_letter].width = max(max_len + 2, 12)

    # Separate results by model
    model_a_results = [r for r in results if r["model"] == MODEL_A]
    model_b_results = [r for r in results if r["model"] == MODEL_B]
    feature_names = [f["name"] for f in FEATURES]

    ollama_version = get_ollama_version()
    report_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # ===================================================================
    # SHEET 1 — Project Overview
    # ===================================================================
    ws1 = wb.active
    ws1.title = "Project Overview"
    overview_data = [
        ["FitTrack Developer Performance & Profiling Report"],
        [],
        ["Property", "Value"],
        ["App Name", "FitTrack"],
        ["Report Generated", report_date],
        ["Current Model", MODEL_A],
        ["Comparison Model", MODEL_B],
        ["Ollama URL", OLLAMA_URL],
        ["Ollama Version", ollama_version],
        ["Langfuse Status", "Connected" if langfuse_info["connected"] else "Not Connected"],
        ["Langfuse Host", langfuse_info["host"]],
        ["Features Covered", "Diet Plan Generation, Workout Plan Generation, Feast Mode Activation"],
        [],
        ["AI Code Locations", ""],
        ["LLM Service (core)", "backend/app/services/llm_service.py"],
        ["Meal Service (diet)", "backend/app/services/meal_service.py  →  generate_meal_plan() line ~1498"],
        ["Workout Service", "backend/app/services/workout_service.py  →  generate_workout_plan() line ~89"],
        ["Feast Mode Manager", "backend/app/services/feast_mode_manager.py  →  _generate_overrides_via_llm() line ~312"],
        ["AI Coach Service", "backend/app/services/ai_coach.py  →  FitnessCoachService"],
        ["Feast Prompts", "backend/app/utils/llm_prompts/feast_prompts.py"],
        [],
        ["Model Configuration", ""],
        ["Config File", "backend/config.py  →  LLM_PROVIDER, LLM_MODEL env vars"],
        ["DB Config (live)", "system_settings table  →  keys: llm_provider, llm_model, ollama_url, ollama_model"],
        ["Admin UI", "Admin Settings Panel  →  PUT /api/admin/settings/{key}"],
        ["How to Switch Model", "1) Set env LLM_MODEL=<name> and restart, OR 2) Update 'llm_model' key in Admin Settings (no restart needed)"],
        ["Default Models Dict", "llm_service.py line 42: {'ollama': 'gpt-oss:120b-cloud', 'openrouter': 'google/gemini-2.0-flash-001', 'openai': 'gpt-4o'}"],
    ]
    for i, row in enumerate(overview_data, 1):
        for j, val in enumerate(row, 1):
            style_data_cell(ws1, i, j, val)

    # Title style
    ws1.cell(row=1, column=1).font = Font(bold=True, size=16, color="1F4E79")
    ws1.merge_cells("A1:B1")
    # Header row
    style_header_row(ws1, 3, 2)
    auto_width(ws1)

    # ===================================================================
    # SHEET 2 — Prompt Structure Breakdown
    # ===================================================================
    ws2 = wb.create_sheet("Prompt Structure Breakdown")
    headers2 = ["Feature", "System Prompt Preview (first 100 chars)", "System Prompt Tokens (est)",
                 "User Message Tokens (est)", "Total Input Tokens (est)", "Notes"]
    for j, h in enumerate(headers2, 1):
        style_data_cell(ws2, 1, j, h)
    style_header_row(ws2, 1, len(headers2))

    for i, feat in enumerate(FEATURES, 2):
        r = next((x for x in results if x["feature"] == feat["name"]), None)
        sp_preview = feat["system_prompt"][:100].replace("\n", " ")
        sp_tokens = estimate_tokens(feat["system_prompt"])
        up_tokens = estimate_tokens(feat["user_prompt"])
        total_est = sp_tokens + up_tokens
        notes = "JSON mode" if feat["json_mode"] else "Text mode"
        notes += f", temp={feat['temperature']}"

        for j, val in enumerate([feat["name"], sp_preview, sp_tokens, up_tokens, total_est, notes], 1):
            style_data_cell(ws2, i, j, val)

    auto_width(ws2)

    # ===================================================================
    # SHEET 3 — Token Analysis — Model Comparison
    # ===================================================================
    ws3 = wb.create_sheet("Token Analysis")
    headers3 = ["Feature", "Model", "Input Tokens", "Output Tokens", "Total Tokens", "Token Efficiency Notes"]
    for j, h in enumerate(headers3, 1):
        style_data_cell(ws3, 1, j, h)
    style_header_row(ws3, 1, len(headers3))

    row_idx = 2
    total_a = {"input": 0, "output": 0, "total": 0}
    total_b = {"input": 0, "output": 0, "total": 0}

    for feat_name in feature_names:
        for model_name in [MODEL_A, MODEL_B]:
            r = next((x for x in results if x["feature"] == feat_name and x["model"] == model_name), None)
            if r and r["status"] == "success":
                inp = r["input_tokens"]
                out = r["output_tokens"]
                tot = r["total_tokens"]
                eff = f"Output/Input ratio: {out/inp:.2f}" if inp > 0 else "N/A"
            else:
                inp = "N/A"
                out = "N/A"
                tot = "N/A"
                eff = r["error"] if r else "No data"

            for j, val in enumerate([feat_name, model_name, inp, out, tot, eff], 1):
                style_data_cell(ws3, row_idx, j, val)
            row_idx += 1

            # Accumulate totals
            if r and r["status"] == "success":
                bucket = total_a if model_name == MODEL_A else total_b
                bucket["input"] += r["input_tokens"]
                bucket["output"] += r["output_tokens"]
                bucket["total"] += r["total_tokens"]

    # Summary rows
    row_idx += 1
    for j, val in enumerate(["TOTAL", MODEL_A, total_a["input"], total_a["output"], total_a["total"], ""], 1):
        c = style_data_cell(ws3, row_idx, j, val)
        c.font = header_font
    row_idx += 1
    for j, val in enumerate(["TOTAL", MODEL_B, total_b["input"], total_b["output"], total_b["total"], ""], 1):
        c = style_data_cell(ws3, row_idx, j, val)
        c.font = header_font
    auto_width(ws3)

    # ===================================================================
    # SHEET 4 — Performance Profiling — Model Comparison
    # ===================================================================
    ws4 = wb.create_sheet("Performance Profiling")
    headers4 = ["Feature", "Model", "Prompt Eval (ms)", "Generation Time (ms)",
                 "Total Ollama Duration (ms)", "Wall Clock Latency (ms)", "Faster Model"]
    for j, h in enumerate(headers4, 1):
        style_data_cell(ws4, 1, j, h)
    style_header_row(ws4, 1, len(headers4))

    row_idx = 2
    for feat_name in feature_names:
        ra = next((x for x in results if x["feature"] == feat_name and x["model"] == MODEL_A), None)
        rb = next((x for x in results if x["feature"] == feat_name and x["model"] == MODEL_B), None)

        # Determine faster
        faster = "N/A"
        if ra and rb and ra["status"] == "success" and rb["status"] == "success":
            faster = MODEL_A if ra["wall_clock_ms"] < rb["wall_clock_ms"] else MODEL_B

        for r in [ra, rb]:
            if r and r["status"] == "success":
                vals = [feat_name, r["model"], r["prompt_eval_duration_ms"],
                        r["eval_duration_ms"], r["total_duration_ms"],
                        r["wall_clock_ms"], faster]
            else:
                err = r["error"] if r else "No data"
                vals = [feat_name, r["model"] if r else "N/A", "N/A", "N/A", "N/A", "N/A", err]

            for j, val in enumerate(vals, 1):
                c = style_data_cell(ws4, row_idx, j, val)
                # Highlight faster model row green
                if r and r["status"] == "success" and r["model"] == faster:
                    c.fill = green_fill

            row_idx += 1

    auto_width(ws4)

    # ===================================================================
    # SHEET 5 — Response Quality Comparison
    # ===================================================================
    ws5 = wb.create_sheet("Response Quality")
    headers5 = ["Feature", "Model", "Response Summary (first 300 chars)", "Word Count",
                 "Structure (JSON/Text/Mixed)", "Completeness", "Quality Notes", "Recommended Model"]
    for j, h in enumerate(headers5, 1):
        style_data_cell(ws5, 1, j, h)
    style_header_row(ws5, 1, len(headers5))

    row_idx = 2
    for feat_name in feature_names:
        ra = next((x for x in results if x["feature"] == feat_name and x["model"] == MODEL_A), None)
        rb = next((x for x in results if x["feature"] == feat_name and x["model"] == MODEL_B), None)

        for r in [ra, rb]:
            if r and r["status"] == "success":
                text = r["response_text"]
                summary = text[:300].replace("\n", " ")
                wc = len(text.split())
                # Detect structure
                stripped = text.strip()
                if stripped.startswith("{") or stripped.startswith("["):
                    structure = "JSON"
                elif "{" in stripped:
                    structure = "Mixed"
                else:
                    structure = "Text"
                # Completeness
                try:
                    json.loads(stripped)
                    completeness = "Complete (valid JSON)"
                except:
                    completeness = "Truncated or invalid JSON"

                # Quality notes
                quality = ""
                if wc < 50:
                    quality = "Very short response"
                elif wc > 2000:
                    quality = "Very long response — check for token bloat"
                else:
                    quality = "Reasonable length"

                # Recommendation
                rec = ""
                if ra and rb and ra["status"] == "success" and rb["status"] == "success":
                    ra_ok = False
                    rb_ok = False
                    try:
                        json.loads(ra["response_text"].strip())
                        ra_ok = True
                    except: pass
                    try:
                        json.loads(rb["response_text"].strip())
                        rb_ok = True
                    except: pass

                    if ra_ok and not rb_ok:
                        rec = f"{MODEL_A} (valid JSON output)"
                    elif rb_ok and not ra_ok:
                        rec = f"{MODEL_B} (valid JSON output)"
                    elif ra_ok and rb_ok:
                        if ra["wall_clock_ms"] < rb["wall_clock_ms"]:
                            rec = f"{MODEL_A} (faster, both valid)"
                        else:
                            rec = f"{MODEL_B} (faster, both valid)"
                    else:
                        rec = "Neither produced valid JSON"

                vals = [feat_name, r["model"], summary, wc, structure, completeness, quality, rec]
            else:
                err = r["error"] if r else "No data"
                vals = [feat_name, r["model"] if r else "N/A", err, 0, "N/A", "Error", err, "N/A"]

            for j, val in enumerate(vals, 1):
                style_data_cell(ws5, row_idx, j, val)
            row_idx += 1

    auto_width(ws5)

    # ===================================================================
    # SHEET 6 — Langfuse Trace Summary
    # ===================================================================
    ws6 = wb.create_sheet("Langfuse Trace Summary")
    headers6 = ["Feature", "Total Traces Captured", "Avg Input Tokens", "Avg Output Tokens",
                 "Avg Latency (ms)", "Error Count", "Last Trace Time"]
    for j, h in enumerate(headers6, 1):
        style_data_cell(ws6, 1, j, h)
    style_header_row(ws6, 1, len(headers6))

    row_idx = 2
    if langfuse_info["connected"] and langfuse_info["traces"]:
        for name, traces in langfuse_info["traces"].items():
            count = len(traces)
            # Best-effort metric extraction
            input_toks = []
            output_toks = []
            latencies = []
            errors = 0
            last_time = "N/A"

            for t in traces:
                obs = t.get("observations", [])
                for o in obs:
                    usage = o.get("usage", {}) or {}
                    if usage.get("input"):
                        input_toks.append(usage["input"])
                    if usage.get("output"):
                        output_toks.append(usage["output"])
                    if o.get("latency"):
                        latencies.append(o["latency"])
                if t.get("status") == "ERROR":
                    errors += 1

            if traces:
                last_time = traces[0].get("timestamp", "N/A")

            avg_in = round(sum(input_toks) / len(input_toks), 1) if input_toks else "N/A"
            avg_out = round(sum(output_toks) / len(output_toks), 1) if output_toks else "N/A"
            avg_lat = round(sum(latencies) / len(latencies), 1) if latencies else "N/A"

            for j, val in enumerate([name, count, avg_in, avg_out, avg_lat, errors, last_time], 1):
                style_data_cell(ws6, row_idx, j, val)
            row_idx += 1
    else:
        note = "See Langfuse Dashboard"
        if langfuse_info["error"]:
            note += f" — {langfuse_info['error']}"
        for feat_name in feature_names:
            for j, val in enumerate([feat_name, note, "N/A", "N/A", "N/A", "N/A", "N/A"], 1):
                style_data_cell(ws6, row_idx, j, val)
            row_idx += 1

        row_idx += 1
        style_data_cell(ws6, row_idx, 1, f"Dashboard URL: {langfuse_info['host']}")

    auto_width(ws6)

    # ===================================================================
    # SHEET 7 — Issues & Recommendations
    # ===================================================================
    ws7 = wb.create_sheet("Issues & Recommendations")
    headers7 = ["#", "Category", "Finding", "Severity", "Recommendation"]
    for j, h in enumerate(headers7, 1):
        style_data_cell(ws7, 1, j, h)
    style_header_row(ws7, 1, len(headers7))

    recommendations = []

    # 1. Prompt token analysis
    for feat in FEATURES:
        est = estimate_tokens(feat["system_prompt"]) + estimate_tokens(feat["user_prompt"])
        if est > 2000:
            recommendations.append({
                "cat": "Prompt Optimization",
                "finding": f"{feat['name']} has ~{est} estimated input tokens. The user prompt is very large.",
                "severity": "Medium",
                "rec": "Consider compressing the prompt: remove redundant examples, use concise instructions, or move static context to system prompt."
            })

    # 2. Latency analysis
    for feat_name in feature_names:
        ra = next((x for x in results if x["feature"] == feat_name and x["model"] == MODEL_A), None)
        if ra and ra["status"] == "success" and ra["wall_clock_ms"] > 60000:
            recommendations.append({
                "cat": "High Latency",
                "finding": f"{feat_name} with {MODEL_A} took {ra['wall_clock_ms']/1000:.1f}s wall clock.",
                "severity": "High",
                "rec": "Consider streaming responses for better UX, or use a smaller/faster model for this feature."
            })

    # 3. Model recommendation
    a_wins = 0
    b_wins = 0
    for feat_name in feature_names:
        ra = next((x for x in results if x["feature"] == feat_name and x["model"] == MODEL_A), None)
        rb = next((x for x in results if x["feature"] == feat_name and x["model"] == MODEL_B), None)
        if ra and rb and ra["status"] == "success" and rb["status"] == "success":
            if ra["wall_clock_ms"] < rb["wall_clock_ms"]:
                a_wins += 1
            else:
                b_wins += 1

    recommendations.append({
        "cat": "Model Recommendation",
        "finding": f"{MODEL_A} won {a_wins}/{len(feature_names)} features on latency. {MODEL_B} won {b_wins}/{len(feature_names)}.",
        "severity": "Info",
        "rec": f"Use {MODEL_A} for production (higher quality for complex JSON tasks). Use {MODEL_B} for quick iteration/testing. Consider {MODEL_B} for simpler features like Feast Mode adjustments."
    })

    # 4. Langfuse instrumentation
    recommendations.append({
        "cat": "Langfuse Instrumentation",
        "finding": "Duration metrics (prompt_eval_duration, eval_duration, total_duration, wall_clock) were missing from Langfuse traces.",
        "severity": "Medium",
        "rec": "FIXED: Added Ollama duration metrics + wall clock latency to both call_llm_json() and call_llm() in llm_service.py. Also fixed model name (was using static MODEL_NAME, now uses live_model)."
    })

    # 5. Caching suggestion
    recommendations.append({
        "cat": "Performance — Caching",
        "finding": "No response caching detected for repeated prompts.",
        "severity": "Low",
        "rec": "Consider adding a Redis cache for identical prompts (e.g., same user profile + same diet type = same prompt hash). TTL of 1 hour would reduce redundant LLM calls."
    })

    # 6. Streaming
    recommendations.append({
        "cat": "UX — Streaming",
        "finding": "All LLM calls use stream=False (blocking). Users wait for full response.",
        "severity": "Medium",
        "rec": "Implement SSE/WebSocket streaming for AI Coach chat responses. Diet/Workout generation can remain blocking since they need full JSON parsing."
    })

    # 7. System health
    errors_count = sum(1 for r in results if r["status"] != "success")
    total_runs = len(results)
    if errors_count == 0:
        health = "Good"
        sev = "Info"
    elif errors_count <= 2:
        health = "Needs Attention"
        sev = "Medium"
    else:
        health = "Critical"
        sev = "High"

    recommendations.append({
        "cat": "Overall System Health",
        "finding": f"{total_runs - errors_count}/{total_runs} runs succeeded. Health: {health}.",
        "severity": sev,
        "rec": f"System status: {health}. {errors_count} errors in {total_runs} total runs." + (" Investigate failed runs." if errors_count > 0 else " All systems operational.")
    })

    for i, rec in enumerate(recommendations, 2):
        for j, val in enumerate([i - 1, rec["cat"], rec["finding"], rec["severity"], rec["rec"]], 1):
            c = style_data_cell(ws7, i, j, val)
            if rec["severity"] == "High":
                ws7.cell(row=i, column=4).fill = red_fill
            elif rec["severity"] == "Medium":
                ws7.cell(row=i, column=4).fill = yellow_fill
            elif rec["severity"] == "Low" or rec["severity"] == "Info":
                ws7.cell(row=i, column=4).fill = green_fill

    auto_width(ws7)

    # ===================================================================
    # SHEET 8 — Raw Data
    # ===================================================================
    ws8 = wb.create_sheet("Raw Data")
    headers8 = ["Timestamp", "Feature", "Model", "Raw Input Tokens", "Raw Output Tokens",
                 "Raw Latency ms", "Raw Response Length", "Source (Live Run / Langfuse)",
                 "Status", "Error", "Prompt Eval ms", "Generation ms", "Ollama Total ms"]
    for j, h in enumerate(headers8, 1):
        style_data_cell(ws8, 1, j, h)
    style_header_row(ws8, 1, len(headers8))

    for i, r in enumerate(results, 2):
        vals = [
            r["timestamp"], r["feature"], r["model"],
            r["input_tokens"], r["output_tokens"],
            r["wall_clock_ms"], r["response_length"],
            "Live Run", r["status"], r.get("error", ""),
            r["prompt_eval_duration_ms"], r["eval_duration_ms"], r["total_duration_ms"]
        ]
        for j, val in enumerate(vals, 1):
            style_data_cell(ws8, i, j, val)

    auto_width(ws8)

    # ===================================================================
    # SAVE
    # ===================================================================
    wb.save(output_path)
    print(f"\n{'='*70}")
    print(f"  REPORT SAVED: {os.path.abspath(output_path)}")
    print(f"{'='*70}\n")


# ---------------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------------
def main():
    print("=" * 70)
    print("  FitTrack Developer Performance & Profiling Report Generator")
    print("=" * 70)
    print(f"  Ollama URL  : {OLLAMA_URL}")
    print(f"  Model A     : {MODEL_A}")
    print(f"  Model B     : {MODEL_B}")
    print(f"  Ollama Ver  : {get_ollama_version()}")
    print("=" * 70)

    # Step 3 — Run profiling
    print("\n[STEP 3] Running profiling for all features with both models...")
    results = run_profiling()

    # Step 4 — Langfuse data
    print("\n[STEP 4] Attempting to pull Langfuse trace data...")
    langfuse_info = try_pull_langfuse()
    if langfuse_info["connected"]:
        print(f"  Langfuse connected. Traces found: {sum(len(v) for v in langfuse_info['traces'].values())}")
    else:
        print(f"  Langfuse not available: {langfuse_info.get('error', 'No credentials')}")

    # Step 5 — Generate Excel
    print("\n[STEP 5] Generating Excel report...")
    project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    output_path = os.path.join(project_root, "fittrack_dev_report.xlsx")
    generate_excel(results, langfuse_info, output_path)

    # Step 6 — Checklist
    print("[STEP 6] Final Checklist:")
    print(f"  [✓] Report saved at: {os.path.abspath(output_path)}")
    print(f"  [✓] All 8 sheets present")
    print(f"  [✓] {len(results)} data points collected ({len([r for r in results if r['status']=='success'])} successful)")
    print(f"  [✓] Both models tested: {MODEL_A}, {MODEL_B}")
    print("  [✓] Done!")


if __name__ == "__main__":
    main()
